#!/usr/bin/env python3
"""Régénère motsatrouver.js depuis MotPsy_V52_APA.xlsx (onglet Motpsy, lignes 3-138)."""
import openpyxl

XLSX = "MotPsy_V52_APA.xlsx"
SHEET = "Motpsy"
FIRST_ROW = 3
LAST_ROW = 138

COLS = {
    "mot": 5,       # E
    "def": 6,       # F
    "cat": 4,       # D
    "photo": 11,    # K
    "cacher": 12,   # L
    "date": 3,      # C
    "indice": 13,   # M
    "exemple": 7,   # G (pas H)
    "rebonds": 9,   # I (pas J)
}

HEADER = """// Liste des mots Motpsy
// Format : [mot, définition, catégorie, photo, cacher1ereLettre, date, indice, exemple, rebonds]
// Catégories : Classique, VO, VIP, Lacan  (le tiret est détecté automatiquement)
// La date (YYYY-MM-DD) détermine le jour de diffusion. Fallback aléatoire si aucune date ne correspond.
// L'indice (si présent) fait apparaître le bouton 💡 ; vide = pas de bouton.

"""


def echapper(valeur):
    """Échappe une valeur de cellule pour un littéral JS entre apostrophes."""
    if valeur is None:
        return ""
    s = str(valeur)
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    s = s.replace("\\", "\\\\")
    s = s.replace("'", "\\'")
    s = s.replace("\n", "\\n")
    return s


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb[SHEET]

    lignes_js = []
    total = 0
    for r in range(FIRST_ROW, LAST_ROW + 1):
        mot = ws.cell(row=r, column=COLS["mot"]).value
        if mot is None or str(mot).strip() == "":
            continue
        total += 1

        definition = echapper(ws.cell(row=r, column=COLS["def"]).value)
        categorie = echapper(ws.cell(row=r, column=COLS["cat"]).value)
        photo = echapper(ws.cell(row=r, column=COLS["photo"]).value)

        cacher_val = ws.cell(row=r, column=COLS["cacher"]).value
        cacher = isinstance(cacher_val, str) and cacher_val.strip().lower() == "oui"

        date_val = ws.cell(row=r, column=COLS["date"]).value
        date_str = date_val.strftime("%Y-%m-%d")

        indice = echapper(ws.cell(row=r, column=COLS["indice"]).value)
        exemple = echapper(ws.cell(row=r, column=COLS["exemple"]).value)
        rebonds = echapper(ws.cell(row=r, column=COLS["rebonds"]).value)
        mot_echappe = echapper(mot)

        champs = [
            mot_echappe, definition, categorie, photo,
            "true" if cacher else "false", date_str, indice, exemple, rebonds,
        ]
        parties = []
        for i, champ in enumerate(champs):
            if i == 4:
                parties.append(champ)  # true/false, pas de quotes
            else:
                parties.append(f"'{champ}'")
        lignes_js.append("  [" + ", ".join(parties) + "],")

    contenu = HEADER + "const LISTE_MOTS_A_TROUVER = [\n" + "\n".join(lignes_js) + "\n];\n"

    with open("motsatrouver.js.new", "w", encoding="utf-8") as f:
        f.write(contenu)

    print(f"{total} entrées écrites dans motsatrouver.js.new")


if __name__ == "__main__":
    main()
